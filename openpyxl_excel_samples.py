# -*- coding: utf-8 -*-
"""
Created on Sun Sep 13 18:35:53 2020

@author: heart
"""

### Read in two the similar reports and write the results back to second one (can be anhyone) without changing format
 
# import openpyxl
import openpyxl

# from openpyxl import Alignment
from openpyxl.styles import Alignment

### Read in excel files 
wb1 = openpyxl.load_workbook(f"excelfile1.xlsx")
wb2 = openpyxl.load_workbook(f"excelfile2.xlsx")

### List all sheetnames 
wb1.sheetnames
wb2.sheetnames

### List worksheets - 0 means the default first worksheet
ws1 = wb1.worksheets[0]
ws2 = wb2.worksheets[0]

### To read in header from file2 and prepare to write back as Header for cell A1
header1 = ws1['A1'].value
header2 = ws2['A1'].value

header = header2 + " vs\n FFF - For KPI 2020"

### Write the header value back to Excel file 2
ws2['A1'] = header

### Remove any ",", or spaces from the read in value string
ali_student1 = ws1['B2'].value.strip().replace(',','')
ali_student2 = ws2['B2'].value.strip().replace(',','')

### Convert string into integer
stu1 = int(ali_student1)
stu2 = int(ali_student2)

### Calculate changes
b2 = (stu2 - stu1) / stu1

### Format the value as 2 digit decimal percent
cell_b2 = "{:.2%}".format(b2)

### Write this value to file 2 cell B2 or row 2 column 2
ws2.cell(2,2).value = cell_b2

### Write additioanl cell values after process
ws1_b3 = ws1['B3'].value
ws2_b3 = ws2['B3'].value

ws1_d3 = ws1['D3'].value
ws2_d3 = ws2['D3'].value

cell_d3 = ws2_d3 + '\nvs.' + '\n' + ws1_d3

### Different method to write cell value back
ws2['D3'] = cell_d3

ws1_d4 = ws1['D4'].value
ws2_d4 = ws2['D4'].value

cell_d4 = "{:.2%}".format((ws2_d4 - ws1_d4)/ ws1_d4)

ws2['D4'] = cell_d4


### To get the changes year over year for all the rows and columns which are comparable 
for row in range(6, 12):
    for col in range(2, 6):
        v1 = ws1.cell(row, col).value
        v2 = ws2.cell(row, col).value
        
        if isinstance(v1, str):
            v1 = v1.replace(',','').replace('%','')
 
        if isinstance(v2, str):
           v2 = v2.replace(',','').replace('%','')
            
        if col == 2 or col == 4:
            v1 = int(v1)
            v2 = int(v2)
        else:
            v1 = float(v1)
            v2 = float(v2)
            
        change = "{:.2%}".format((v2 - v1)/ v1)
        ws2.cell(row, col).value = change
        
ws2.row_dimensions[1].height = 100
ws2.row_dimensions[3].height = 50
al = Alignment(horizonal = 'center', vertical = 'center')
ws2.cell(2, 2).alignment= al

wb2.save("excelfile2.xlsx")
